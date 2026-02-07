#!/usr/bin/env python3
"""
Transform Startups.watch Excel export to BuildAtlas CSV format.

Usage:
  python3 scripts/transform-startups-watch.py <input.xlsx> [output.csv]

If output is not specified, writes to apps/web/data/tr/2026-02/input/startups.csv
"""

import csv
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print(
        "ERROR: Missing dependency 'openpyxl'.\n"
        "Install it first, then re-run:\n"
        "  python3 -m pip install openpyxl\n",
        file=sys.stderr,
    )
    raise SystemExit(2)


# Funding type mapping: Startups.watch → BuildAtlas
FUNDING_TYPE_MAP = {
    "Pre Seed": "Pre-Seed",
    "Seed": "Seed",
    "Post Seed": "Post-Seed",
    "Series A": "Series A",
    "Series B": "Series B",
    "Series C": "Series C",
    "Series D": "Series D",
    "Series E": "Series E",
    "Safe": "SAFE",
    "Bridge": "Bridge",
    "Convertible Note": "Convertible Note",
    "Equity Crowdfunding": "Equity Crowdfunding",
    "Corporate Round": "Corporate Round",
}

# Funding stage mapping (broader categories for BuildAtlas)
FUNDING_STAGE_MAP = {
    "Pre Seed": "Pre-Seed",
    "Seed": "Seed",
    "Post Seed": "Seed",
    "Safe": "Pre-Seed",
    "Bridge": "Early Stage Venture",
    "Convertible Note": "Pre-Seed",
    "Equity Crowdfunding": "Seed",
    "Corporate Round": "Late Stage Venture",
    "Series A": "Early Stage Venture",
    "Series B": "Early Stage Venture",
    "Series C": "Late Stage Venture",
    "Series D": "Late Stage Venture",
    "Series E": "Late Stage Venture",
}

# BuildAtlas CSV headers
CSV_HEADERS = [
    "Transaction Name",
    "Transaction Name URL",
    "Funding Type",
    "Money Raised",
    "Money Raised Currency",
    "Money Raised (in USD)",
    "Announced Date",
    "Funding Stage",
    "Organization Description",
    "Organization Website",
    "Number of Funding Rounds",
    "Organization Industries",
    "Organization Location",
    "Lead Investors",
]


def parse_funding_amount(value: str | None) -> str:
    """Parse '$21,150' or '$1,000,000' to '21150' or '1000000'."""
    if not value:
        return ""
    # Remove $ and commas, strip whitespace
    cleaned = str(value).replace("$", "").replace(",", "").strip()
    # Handle float strings like '21150.0'
    try:
        num = float(cleaned)
        if num > 0:
            return str(int(num))
    except (ValueError, TypeError):
        pass
    return ""


def format_location(headquarter: str | None) -> str:
    """Convert 'Istanbul / Turkey' to 'Istanbul, Turkey, Asia'."""
    if not headquarter:
        return ""
    # Split on ' / ' separator
    parts = [p.strip() for p in str(headquarter).split("/")]
    city = parts[0] if len(parts) > 0 else ""
    country = parts[1] if len(parts) > 1 else "Turkey"
    # All Turkish cities map to Asia continent
    return f"{city}, {country}, Asia"


def format_announced_date(year: int | str | None) -> str:
    """Convert year (2026) to date string (2026-01-01)."""
    if not year:
        return ""
    try:
        y = int(year)
        return f"{y}-01-01"
    except (ValueError, TypeError):
        return ""


def transform(input_path: str, output_path: str) -> None:
    """Transform Startups.watch Excel to BuildAtlas CSV."""
    wb = load_workbook(input_path, read_only=True, data_only=True)
    ws = wb["startups"]

    # Read headers
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {h: i for i, h in enumerate(headers) if h}

    def cell(row, col_name):
        idx = header_map.get(col_name)
        if idx is None:
            return None
        val = row[idx].value if row[idx] else None
        return str(val).strip() if val is not None else None

    rows_written = 0
    skipped = 0

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_HEADERS)
        writer.writeheader()

        for row in ws.iter_rows(min_row=2):
            startup_name = cell(row, "Startup")
            if not startup_name:
                skipped += 1
                continue

            funding_type_raw = cell(row, "Last Funding Type") or ""
            funding_amount = parse_funding_amount(cell(row, "Last Funding Amount"))

            writer.writerow({
                "Transaction Name": startup_name,
                "Transaction Name URL": cell(row, "Links") or "",
                "Funding Type": FUNDING_TYPE_MAP.get(funding_type_raw, funding_type_raw),
                "Money Raised": funding_amount,
                "Money Raised Currency": "USD" if funding_amount else "",
                "Money Raised (in USD)": funding_amount,
                "Announced Date": format_announced_date(cell(row, "Year")),
                "Funding Stage": FUNDING_STAGE_MAP.get(funding_type_raw, ""),
                "Organization Description": cell(row, "Description") or "",
                "Organization Website": cell(row, "Links") or "",
                "Number of Funding Rounds": "1" if funding_amount else "",
                "Organization Industries": cell(row, "Category") or "",
                "Organization Location": format_location(cell(row, "Headquarter")),
                "Lead Investors": "",
            })
            rows_written += 1

    wb.close()
    print(f"Transformed {rows_written} startups ({skipped} skipped)")
    print(f"Output: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(f"Usage: {sys.argv[0]} <input.xlsx> [output.csv]")
        sys.exit(1)

    input_file = sys.argv[1]
    default_output = str(
        Path(__file__).parent.parent
        / "apps/web/data/tr/2026-02/input/startups.csv"
    )
    output_file = sys.argv[2] if len(sys.argv) > 2 else default_output

    # Ensure output directory exists
    Path(output_file).parent.mkdir(parents=True, exist_ok=True)

    transform(input_file, output_file)
